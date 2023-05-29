import openpyxl, json
import openai

# pyinstaller -F -n TranslationByChatGPT main.py

def Init():
    print("프로그램 개발자 : 갈대(westreed@naver.com)\n")
    print("* OpenAI를 사용하기 위해, key.json에서 apiKey를 확인합니다...")
    try:
        with open("key.json", "r") as f:
            api_key = json.load(f)["api_key"].strip()
            if api_key in ['', None]:
                print("* apiKey를 확인하지 못했습니다!\nkey.json의 'api_key' : '여기'에서 OpenAI에서 발급한 apiKey를 넣어주세요.")
                return False
            openai.api_key = api_key
            print("* apiKey를 확인했습니다.\n")
    except:
        print("* key.json 파일이 없습니다! 파일을 다시 생성합니다.")
        with open("key.json", "w") as f:
            f.write('{\n\t"api_key" : ""\n}')
        return False
    
    model_engine = "gpt-3.5-turbo"
    print(f"* 번역을 도와줄 엔진은 `{model_engine}`입니다.\n")

    filename = input("번역대상 파일의 이름을 입력해주세요 : ")
    print(f"* {filename}.xlsx 파일을 여는 중입니다.")
    if filename.startswith(".xlsx"): filename.replace(".xlsx", "")

    wb = None
    ws = None
    try:
        wb = openpyxl.load_workbook(f'{filename}.xlsx',data_only=True)
        ws = wb.active
        print(f"* {filename}.xlsx 파일을 열었습니다.\n")
    except:
        print(f"* {filename}.xlsx 파일이 없습니다!")
        return False

    print(f"* 프롬포트 명령어는 {filename}.xlsx의 가장 첫번째줄을 기준으로 합니다.")
    prompt = ws.cell(row=1, column=1).value
    if prompt is not None:
        print(f"* 프롬포트 명령어는 `{prompt}` 입니다.")
        prompt = [{"role":"system", "content":prompt}]
    else:
        default_text = "당신은 번역가입니다. 주어진 문장을 한글로 자연스럽게 번역해야 합니다. 부연설명은 따로 필요없습니다. 또한, 의성어도 그대로 번역해주세요."
        print(f"* 프롬포트 명령어는 기본값으로 설정된 `{default_text}` 입니다.")
        prompt = [{"role":"system", "content":default_text}]
        

    cnt = 1
    while True:
        cnt += 1
        c = ws.cell(row=cnt, column=1)
        if c.value is None: break

    if cnt > 2:
        cnt -= 2
        print(f"* 총 문장수는 {cnt}개 입니다. 지금부터 번역을 시작합니다.\n")
    else:
        print(f"* 번역할 문장이 없습니다!")
        return False

    for idx in range(cnt):
        row = idx + 2
        c = ws.cell(row=row, column=1).value
        _prompt = prompt + [{"role":"user", "content":f"{c}"}]
        try:
            competion = openai.ChatCompletion.create(
                model=model_engine,
                messages=_prompt
            )
            res = competion['choices'][0]['message']['content']
            ws.cell(row=row, column=2, value=res)
            print(f"* ({idx+1:04}/{cnt:04}) {c} -> {res}")
        except:
            ws.cell(row=row, column=2, value="Error")
            print(f"* ({idx+1:04}/{cnt:04}) {c} -> Error 발생")

    wb.save('result.xlsx')
    print("\n* 모든 문장의 번역이 끝났습니다.")
    input("* 종료하려면 엔터키를 눌러주세요...")

if __name__ == "__main__":
    Init()